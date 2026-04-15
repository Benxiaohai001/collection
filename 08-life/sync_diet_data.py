#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 同步减肥日记数据：从 md 文件增量更新到 CSV 和 Excel

from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import os

BASE_DIR = "/Users/baker/Documents/code/collection/08-life"
MD_FILE = os.path.join(BASE_DIR, "小胖子减肥日记.md")
CSV_FILE = os.path.join(BASE_DIR, "小胖子减肥日记.csv")
EXCEL_FILE = os.path.join(BASE_DIR, "小胖子减肥日记.xlsx")

weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

def parse_md_file():
    records = []
    with open(MD_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) >= 2:
                date_str = parts[0]
                weight = parts[1]
                try:
                    date_obj = datetime.strptime(date_str, "%Y%m%d")
                    records.append({
                        'date': date_obj.strftime("%Y-%m-%d"),
                        'weekday': weekdays[date_obj.weekday()],
                        'weight': float(weight)
                    })
                except ValueError:
                    continue
    return records

def get_existing_csv_dates():
    dates = set()
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                dates.add(row['日期'])
    return dates

def get_existing_excel_dates():
    dates = set()
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                dates.add(str(row[0]))
        wb.close()
    return dates

def init_csv():
    with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['日期', '周几', '体重(kg)', '备注'])

def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "减肥日记"
    ws.append(["日期", "周几", "体重(kg)", "备注"])
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    wb.save(EXCEL_FILE)

def append_to_csv(records):
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for r in records:
            writer.writerow([r['date'], r['weekday'], r['weight'], ''])

def append_to_excel(records):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for r in records:
        ws.append([r['date'], r['weekday'], r['weight'], ''])
    wb.save(EXCEL_FILE)
    wb.close()

def sync():
    all_records = parse_md_file()
    csv_dates = get_existing_csv_dates()
    excel_dates = get_existing_excel_dates()
    
    new_for_csv = [r for r in all_records if r['date'] not in csv_dates]
    new_for_excel = [r for r in all_records if r['date'] not in excel_dates]
    
    if not os.path.exists(CSV_FILE):
        init_csv()
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    
    if new_for_csv:
        append_to_csv(new_for_csv)
        print(f"CSV 新增 {len(new_for_csv)} 条记录")
    
    if new_for_excel:
        append_to_excel(new_for_excel)
        print(f"Excel 新增 {len(new_for_excel)} 条记录")
    
    if not new_for_csv and not new_for_excel:
        print("无新增记录")

if __name__ == "__main__":
    sync()